import openpyxl
import sys
import pyperclip
import webbrowser

FORM_TAB        = 'Form responses 1'
SELECTED_TAB    = 'selected_caditates'
WAIT_LIST_TAB   = 'waiting_list'

SELECT_MSG_FILE = 'selected_msg.txt'
WAIT_MSG_FILE   = 'waiting_msg.txt'

WHATSAPP_URL    = 'https://web.whatsapp.com/send?phone={}&text&type=phone_number&app_absent=0'


def get_form_students_details(ws):
    details = []

    maximum_row = ws.max_row

    for row in range(2, maximum_row + 1):

        name = ws[f'C{row}'].value
        if name == None:
            name = ''

        email = ws[f'B{row}'].value
        if email == None:
            email = ''

        phone = ws[f'G{row}'].value
        if phone == None:
            phone = ''

        if name:
            details.append([name, email, phone])

    return details


def get_student_names(ws, name_col):
    students = []

    maximum_row = ws.max_row

    for row in range(2, maximum_row + 1):

        name = ws[f'{name_col}{row}'].value
        if name:
            students.append(name)

    return students


def filter_students_details(students, details):
    filter_details = []
    for student in students:
        for i in range(len(details)):
            if (student == details[i][0]):
                filter_details.append(details[i])
                break
    return filter_details


def get_file_utf8_message(file_name):
    text_contents = ''
    with open(file_name, 'rb') as filein:
        contents = filein.read()
        text_contents = contents.decode('utf-8', 'ignore')
    return text_contents


if __name__ == '__main__':
    if (len(sys.argv) != 2):
        print(f'Argument error\n Usage {sys.argv[0]} <excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    details = get_form_students_details(wb[FORM_TAB])

    students = get_student_names(wb[SELECTED_TAB], 'A')
    selected_students = filter_students_details(students, details)

    students = get_student_names(wb[WAIT_LIST_TAB], 'A')
    waiting_students = filter_students_details(students, details)

    selected_msg = get_file_utf8_message(SELECT_MSG_FILE)
    waiting_msg = get_file_utf8_message(WAIT_MSG_FILE)

    print(f'Selected students ({len(selected_students)}):')
    for [name, email, phone] in selected_students:
        print(f'{name:32} {email:64} {phone:10}')
        msg = f'Dear {name},\n\n{selected_msg}'
        pyperclip.copy(msg)
        phone_number = f'+94{phone}'
        url = WHATSAPP_URL.format(phone_number)
        webbrowser.open_new_tab(url)
        input('Enter to continue ...')

    print(f'\n\nWaiting list students ({len(waiting_students)}):')
    for [name, email, phone] in waiting_students:
        print(f'{name:32} {email:64} {phone:10}')
        msg = f'Dear {name},\n\n{waiting_msg}'
        pyperclip.copy(msg)
        phone_number = f'+94{phone}'
        url = WHATSAPP_URL.format(phone_number)
        webbrowser.open_new_tab(url)
        input('Enter to continue ...')
